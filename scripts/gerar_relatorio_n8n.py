import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_FUNCIONARIOS_FILE = "funcionarios_json.xlsx"
DEFAULT_PRODUTOS_FILE = "produtos_json.xlsx"
DEFAULT_TEMPLATE_FILE = "modelo.xlsx"
DEFAULT_OUTPUT_FILE = "relatorio_final.xlsx"


def _copy_row_style(worksheet, source_row, target_row, columns):
    source_dimensions = worksheet.row_dimensions[source_row]
    target_dimensions = worksheet.row_dimensions[target_row]

    if source_dimensions.height is not None:
        target_dimensions.height = source_dimensions.height

    for column in columns:
        source_cell = worksheet[f"{column}{source_row}"]
        target_cell = worksheet[f"{column}{target_row}"]
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def _extend_worksheet_columns(worksheet, total_columns):
    original_max_column = worksheet.max_column
    if total_columns <= original_max_column:
        return

    source_column = get_column_letter(original_max_column)
    source_dimension = worksheet.column_dimensions[source_column]

    for column_index in range(original_max_column + 1, total_columns + 1):
        target_column = get_column_letter(column_index)
        target_dimension = worksheet.column_dimensions[target_column]

        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.bestFit = source_dimension.bestFit
        target_dimension.outlineLevel = source_dimension.outlineLevel

        for row_number in range(1, worksheet.max_row + 1):
            source_cell = worksheet.cell(row=row_number, column=original_max_column)
            target_cell = worksheet.cell(row=row_number, column=column_index)
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)


def _load_source_table(source_file):
    source_path = Path(source_file)
    if not source_path.exists():
        raise ValueError("Arquivo de origem nao encontrado: {}".format(source_file))

    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
    if not any(header is not None for header in headers):
        raise ValueError("Nenhum cabecalho foi encontrado em {}".format(source_file))

    rows = []
    for row_number in range(2, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_number, column=col).value
            for col in range(1, worksheet.max_column + 1)
        ]

        if all(value is None for value in row_values):
            continue

        rows.append(row_values)

    if not rows:
        raise ValueError("Nenhuma linha de dados foi encontrada em {}".format(source_file))

    return headers, rows


def _fill_worksheet(worksheet, headers, source_rows):
    header_row = 1
    start_row = 2
    style_row = 2
    original_max_row = worksheet.max_row
    total_columns = len(headers)
    _extend_worksheet_columns(worksheet, total_columns)

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=header_row, column=column_index, value=header)

    for row_number, source_row in enumerate(source_rows, start=start_row):
        if row_number > original_max_row:
            destination_columns = tuple(
                worksheet.cell(row=style_row, column=column_index).column_letter
                for column_index in range(1, total_columns + 1)
            )
            _copy_row_style(worksheet, style_row, row_number, destination_columns)

        for column_index, value in enumerate(source_row, start=1):
            worksheet.cell(row=row_number, column=column_index, value=value)


def _fill_template(tables, template_file, output_file):
    template_path = Path(template_file)
    if not template_path.exists():
        raise ValueError("Arquivo de modelo nao encontrado: {}".format(template_file))

    workbook = load_workbook(template_path)
    template_worksheet = workbook[workbook.sheetnames[0]]
    worksheets = [template_worksheet]

    for _ in tables[1:]:
        worksheets.append(workbook.copy_worksheet(template_worksheet))

    for worksheet, (sheet_name, headers, source_rows) in zip(worksheets, tables):
        worksheet.title = sheet_name
        _fill_worksheet(worksheet, headers, source_rows)

    workbook.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Copia os dados de funcionarios_json.xlsx e produtos_json.xlsx "
            "para abas separadas de modelo.xlsx, preservando a formatacao."
        )
    )
    parser.add_argument(
        "--funcionarios-file",
        default=DEFAULT_FUNCIONARIOS_FILE,
        help="Planilha de funcionarios. Padrao: {}".format(DEFAULT_FUNCIONARIOS_FILE),
    )
    parser.add_argument(
        "--produtos-file",
        default=DEFAULT_PRODUTOS_FILE,
        help="Planilha de produtos. Padrao: {}".format(DEFAULT_PRODUTOS_FILE),
    )
    parser.add_argument(
        "--template-file",
        default=DEFAULT_TEMPLATE_FILE,
        help="Planilha modelo com o design. Padrao: {}".format(DEFAULT_TEMPLATE_FILE),
    )
    parser.add_argument(
        "--output-file",
        default=DEFAULT_OUTPUT_FILE,
        help="Arquivo final gerado. Padrao: {}".format(DEFAULT_OUTPUT_FILE),
    )
    args = parser.parse_args()

    funcionarios_headers, funcionarios_rows = _load_source_table(args.funcionarios_file)
    produtos_headers, produtos_rows = _load_source_table(args.produtos_file)

    tables = [
        ("Funcionarios", funcionarios_headers, funcionarios_rows),
        ("Produtos", produtos_headers, produtos_rows),
    ]
    _fill_template(tables, args.template_file, args.output_file)

    print(
        "Arquivo gerado com sucesso: {} "
        "({} funcionarios e {} produtos)".format(
            args.output_file, len(funcionarios_rows), len(produtos_rows)
        )
    )


if __name__ == "__main__":
    main()
